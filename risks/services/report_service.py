"""
Report generation service for Risk Management Application.
Generates PDF and Excel reports with analytics and charts.
"""

import io
from datetime import datetime
from django.db.models import Count, Avg, Q

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
    from reportlab.graphics.charts.legends import Legend
    from reportlab.graphics.widgets.markers import makeMarker
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.image import Image as OpenpyxlImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import plotly.graph_objects as go
    import plotly.io as pio
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

import pandas as pd
from risks.models import Risk


class ReportService:
    """Service for generating risk management reports."""
    
    # Chart colors
    CHART_COLORS = [
        colors.HexColor('#0d9488'),  # Sea green
        colors.HexColor('#14b8a6'),  # Teal
        colors.HexColor('#2dd4bf'),  # Light teal
        colors.HexColor('#5eead4'),  # Lighter teal
        colors.HexColor('#99f6e4'),  # Very light teal
    ]
    
    SEVERITY_COLORS = [
        colors.HexColor('#ef4444'),  # Red - Critical
        colors.HexColor('#f97316'),  # Orange - High
        colors.HexColor('#eab308'),  # Yellow - Medium
        colors.HexColor('#22c55e'),  # Green - Low
    ]
    
    def __init__(self):
        self.generated_at = datetime.now()
    
    def get_statistics(self):
        """Get all statistics for the report."""
        risks = Risk.objects.all()
        total = risks.count()
        
        if total == 0:
            return {
                'total_risks': 0,
                'by_severity': {},
                'by_status': {},
                'by_category': {},
                'by_owner': {},
                'by_effectiveness': {},
                'mitigated_risks': [],
                'not_mitigated_risks': [],
                'average_score': 0,
            }
        
        # Severity counts
        by_severity = {
            'Critical': risks.filter(risk_score__gte=15).count(),
            'High': risks.filter(risk_score__gte=8, risk_score__lt=15).count(),
            'Medium': risks.filter(risk_score__gte=4, risk_score__lt=8).count(),
            'Low': risks.filter(risk_score__lt=4).count(),
        }
        
        # Status counts
        by_status = dict(risks.values('status').annotate(count=Count('id')).values_list('status', 'count'))
        
        # Category counts
        by_category = dict(risks.values('risk_category').annotate(count=Count('id')).values_list('risk_category', 'count'))
        
        # Owner counts
        by_owner = dict(risks.values('risk_owner').annotate(count=Count('id')).values_list('risk_owner', 'count'))
        
        # Effectiveness counts
        by_effectiveness = dict(risks.values('control_effectiveness').annotate(count=Count('id')).values_list('control_effectiveness', 'count'))
        
        # Mitigated vs Not Mitigated
        mitigated_risks = list(risks.filter(
            Q(is_mitigated=True) | Q(status='Mitigated') | Q(status='Closed')
        ).values('risk_id', 'title', 'risk_score', 'status', 'risk_owner', 'risk_category'))
        
        not_mitigated_risks = list(risks.filter(
            is_mitigated=False
        ).exclude(
            status__in=['Mitigated', 'Closed']
        ).values('risk_id', 'title', 'risk_score', 'status', 'risk_owner', 'risk_category'))
        
        # Average score
        avg_score = risks.aggregate(Avg('risk_score'))['risk_score__avg'] or 0
        
        return {
            'total_risks': total,
            'by_severity': by_severity,
            'by_status': by_status,
            'by_category': by_category,
            'by_owner': by_owner,
            'by_effectiveness': by_effectiveness,
            'mitigated_risks': mitigated_risks,
            'not_mitigated_risks': not_mitigated_risks,
            'average_score': round(avg_score, 2),
        }
    
    def _create_pie_chart(self, data, title, width=280, height=200):
        """Create a pie chart for PDF report."""
        if not data or sum(data.values()) == 0:
            return None
        
        drawing = Drawing(width, height)
        
        pie = Pie()
        pie.x = 80
        pie.y = 30
        pie.width = 120
        pie.height = 120
        pie.data = list(data.values())
        pie.labels = list(data.keys())
        pie.slices.strokeWidth = 0.5
        pie.slices.strokeColor = colors.white
        
        # Apply colors
        for i in range(len(data)):
            pie.slices[i].fillColor = self.CHART_COLORS[i % len(self.CHART_COLORS)]
        
        # Add legend
        legend = Legend()
        legend.x = 220
        legend.y = height - 50
        legend.dx = 8
        legend.dy = 8
        legend.fontName = 'Helvetica'
        legend.fontSize = 8
        legend.boxAnchor = 'nw'
        legend.columnMaximum = 5
        legend.strokeWidth = 0.5
        legend.strokeColor = colors.HexColor('#334155')
        legend.deltax = 60
        legend.deltay = 10
        legend.autoXPadding = 5
        legend.yGap = 0
        legend.dxTextSpace = 5
        legend.alignment = 'right'
        legend.dividerLines = 1|2|4
        legend.subCols.rpad = 30
        
        legend.colorNamePairs = [(self.CHART_COLORS[i % len(self.CHART_COLORS)], 
                                  f'{k}: {v}') for i, (k, v) in enumerate(data.items())]
        
        # Title
        title_text = String(width/2, height - 15, title,
                           fontName='Helvetica-Bold', fontSize=11,
                           fillColor=colors.HexColor('#0d9488'),
                           textAnchor='middle')
        
        drawing.add(pie)
        drawing.add(legend)
        drawing.add(title_text)
        
        return drawing
    
    def _create_severity_pie_chart(self, data, width=280, height=200):
        """Create a pie chart specifically for severity with custom colors."""
        if not data or sum(data.values()) == 0:
            return None
        
        drawing = Drawing(width, height)
        
        pie = Pie()
        pie.x = 80
        pie.y = 30
        pie.width = 120
        pie.height = 120
        pie.data = list(data.values())
        pie.labels = list(data.keys())
        pie.slices.strokeWidth = 0.5
        pie.slices.strokeColor = colors.white
        
        # Apply severity-specific colors
        severity_color_map = {
            'Critical': colors.HexColor('#ef4444'),
            'High': colors.HexColor('#f97316'),
            'Medium': colors.HexColor('#eab308'),
            'Low': colors.HexColor('#22c55e'),
        }
        
        for i, key in enumerate(data.keys()):
            pie.slices[i].fillColor = severity_color_map.get(key, self.CHART_COLORS[i % len(self.CHART_COLORS)])
        
        # Legend
        legend = Legend()
        legend.x = 220
        legend.y = height - 50
        legend.dx = 8
        legend.dy = 8
        legend.fontName = 'Helvetica'
        legend.fontSize = 8
        legend.boxAnchor = 'nw'
        legend.columnMaximum = 5
        legend.strokeWidth = 0.5
        legend.strokeColor = colors.HexColor('#334155')
        legend.deltax = 60
        legend.deltay = 10
        legend.autoXPadding = 5
        legend.dxTextSpace = 5
        legend.alignment = 'right'
        
        legend.colorNamePairs = [(severity_color_map.get(k, self.CHART_COLORS[i % len(self.CHART_COLORS)]), 
                                  f'{k}: {v}') for i, (k, v) in enumerate(data.items())]
        
        # Title
        title_text = String(width/2, height - 15, 'Risk Distribution by Severity',
                           fontName='Helvetica-Bold', fontSize=11,
                           fillColor=colors.HexColor('#0d9488'),
                           textAnchor='middle')
        
        drawing.add(pie)
        drawing.add(legend)
        drawing.add(title_text)
        
        return drawing
    
    def _create_bar_chart(self, data, title, width=450, height=200):
        """Create a horizontal bar chart for PDF report."""
        if not data or len(data) == 0:
            return None
        
        drawing = Drawing(width, height)
        
        chart = HorizontalBarChart()
        chart.x = 120
        chart.y = 30
        chart.width = width - 160
        chart.height = height - 60
        chart.data = [list(data.values())]
        chart.categoryAxis.categoryNames = list(data.keys())
        chart.categoryAxis.labels.fontName = 'Helvetica'
        chart.categoryAxis.labels.fontSize = 8
        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = max(data.values()) * 1.2 if data.values() else 10
        chart.valueAxis.valueStep = max(1, int(max(data.values()) / 5)) if data.values() else 2
        chart.bars[0].fillColor = colors.HexColor('#0d9488')
        chart.bars[0].strokeColor = colors.HexColor('#0f766e')
        chart.bars[0].strokeWidth = 1
        chart.barWidth = 12
        chart.groupSpacing = 8
        
        # Title
        title_text = String(width/2, height - 10, title,
                           fontName='Helvetica-Bold', fontSize=11,
                           fillColor=colors.HexColor('#0d9488'),
                           textAnchor='middle')
        
        drawing.add(chart)
        drawing.add(title_text)
        
        return drawing
    
    def generate_pdf_report(self):
        """Generate a PDF report with analytics and charts."""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF generation. Install with: pip install reportlab")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#0d9488'),
            alignment=1  # Center
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#14b8a6'),
        )
        
        # Get statistics
        stats = self.get_statistics()
        
        # Build document
        elements = []
        
        # Title
        elements.append(Paragraph("Risk Management Report", title_style))
        elements.append(Paragraph(f"Generated on: {self.generated_at.strftime('%B %d, %Y at %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Executive Summary Table
        elements.append(Paragraph("Executive Summary", heading_style))
        summary_data = [
            ['Metric', 'Value'],
            ['Total Risks', str(stats['total_risks'])],
            ['Average Risk Score', str(stats['average_score'])],
            ['Critical Risks', str(stats['by_severity'].get('Critical', 0))],
            ['Mitigated Risks', str(len(stats['mitigated_risks']))],
            ['Risks Requiring Attention', str(len(stats['not_mitigated_risks']))],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d9488')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdfa')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#99f6e4')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Charts Section
        elements.append(Paragraph("Risk Analytics Charts", heading_style))
        elements.append(Spacer(1, 10))
        
        # Severity Pie Chart
        severity_chart = self._create_severity_pie_chart(stats['by_severity'])
        if severity_chart:
            elements.append(severity_chart)
            elements.append(Spacer(1, 20))
        
        # Status Pie Chart
        status_chart = self._create_pie_chart(stats['by_status'], 'Risk Distribution by Status')
        if status_chart:
            elements.append(status_chart)
            elements.append(Spacer(1, 20))
        
        # Category Bar Chart
        if stats['by_category']:
            category_chart = self._create_bar_chart(stats['by_category'], 'Risks by Category')
            if category_chart:
                elements.append(category_chart)
                elements.append(Spacer(1, 20))
        
        # Owner Bar Chart
        if stats['by_owner']:
            owner_chart = self._create_bar_chart(stats['by_owner'], 'Risks by Owner')
            if owner_chart:
                elements.append(owner_chart)
                elements.append(Spacer(1, 20))
        
        # Page break before tables
        elements.append(PageBreak())
        
        # Risk by Severity Table
        elements.append(Paragraph("Risk Distribution by Severity", heading_style))
        severity_data = [['Severity Level', 'Count']]
        for level, count in stats['by_severity'].items():
            severity_data.append([level, str(count)])
        
        severity_table = Table(severity_data, colWidths=[3*inch, 2*inch])
        severity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d9488')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ccfbf1')),
            ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#fee2e2')),
            ('BACKGROUND', (0, 2), (0, 2), colors.HexColor('#ffedd5')),
            ('BACKGROUND', (0, 3), (0, 3), colors.HexColor('#fef3c7')),
            ('BACKGROUND', (0, 4), (0, 4), colors.HexColor('#dcfce7')),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        elements.append(severity_table)
        elements.append(Spacer(1, 20))
        
        # Risk by Status Table
        elements.append(Paragraph("Risk Distribution by Status", heading_style))
        status_data = [['Status', 'Count']]
        for status, count in stats['by_status'].items():
            status_data.append([status, str(count)])
        
        status_table = Table(status_data, colWidths=[3*inch, 2*inch])
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d9488')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ccfbf1')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdfa')),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 30))
        
        # Mitigated Risks Section
        elements.append(Paragraph(f"Mitigated Risks ({len(stats['mitigated_risks'])})", heading_style))
        if stats['mitigated_risks']:
            mitigated_data = [['Risk ID', 'Title', 'Score', 'Status', 'Owner']]
            for risk in stats['mitigated_risks'][:20]:
                mitigated_data.append([
                    risk['risk_id'],
                    risk['title'][:30] + '...' if len(risk['title']) > 30 else risk['title'],
                    str(risk['risk_score']),
                    risk['status'],
                    risk['risk_owner']
                ])
            
            mitigated_table = Table(mitigated_data, colWidths=[0.8*inch, 2.2*inch, 0.6*inch, 0.9*inch, 0.9*inch])
            mitigated_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22c55e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bbf7d0')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdf4')),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))
            elements.append(mitigated_table)
        else:
            elements.append(Paragraph("No mitigated risks found.", styles['Normal']))
        
        elements.append(Spacer(1, 30))
        
        # Not Mitigated Risks Section
        elements.append(Paragraph(f"Risks Requiring Attention ({len(stats['not_mitigated_risks'])})", heading_style))
        if stats['not_mitigated_risks']:
            not_mitigated_data = [['Risk ID', 'Title', 'Score', 'Status', 'Owner']]
            for risk in stats['not_mitigated_risks'][:20]:
                not_mitigated_data.append([
                    risk['risk_id'],
                    risk['title'][:30] + '...' if len(risk['title']) > 30 else risk['title'],
                    str(risk['risk_score']),
                    risk['status'],
                    risk['risk_owner']
                ])
            
            not_mitigated_table = Table(not_mitigated_data, colWidths=[0.8*inch, 2.2*inch, 0.6*inch, 0.9*inch, 0.9*inch])
            not_mitigated_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#fecaca')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fef2f2')),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))
            elements.append(not_mitigated_table)
        else:
            elements.append(Paragraph("All risks have been mitigated!", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generate_csv_report(self):
        """Generate a CSV report with all risk data and summary."""
        import csv
        from io import StringIO
        
        # Get statistics
        stats = self.get_statistics()
        
        # Create string buffer for CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # Write report header
        writer.writerow(['Risk Management Report'])
        writer.writerow([f'Generated on: {self.generated_at.strftime("%B %d, %Y at %H:%M")}'])
        writer.writerow([])
        
        # Write summary section
        writer.writerow(['=== EXECUTIVE SUMMARY ==='])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Risks', stats['total_risks']])
        writer.writerow(['Average Risk Score', stats['average_score']])
        writer.writerow(['Critical Risks', stats['by_severity'].get('Critical', 0)])
        writer.writerow(['High Risks', stats['by_severity'].get('High', 0)])
        writer.writerow(['Medium Risks', stats['by_severity'].get('Medium', 0)])
        writer.writerow(['Low Risks', stats['by_severity'].get('Low', 0)])
        writer.writerow(['Mitigated Risks', len(stats['mitigated_risks'])])
        writer.writerow(['Risks Requiring Attention', len(stats['not_mitigated_risks'])])
        writer.writerow([])
        
        # Write severity breakdown
        writer.writerow(['=== RISK DISTRIBUTION BY SEVERITY ==='])
        writer.writerow(['Severity', 'Count'])
        for sev, count in stats['by_severity'].items():
            writer.writerow([sev, count])
        writer.writerow([])
        
        # Write status breakdown
        writer.writerow(['=== RISK DISTRIBUTION BY STATUS ==='])
        writer.writerow(['Status', 'Count'])
        for status, count in stats['by_status'].items():
            writer.writerow([status, count])
        writer.writerow([])
        
        # Write category breakdown
        writer.writerow(['=== RISK DISTRIBUTION BY CATEGORY ==='])
        writer.writerow(['Category', 'Count'])
        for cat, count in stats['by_category'].items():
            writer.writerow([cat, count])
        writer.writerow([])
        
        # Write owner breakdown
        writer.writerow(['=== RISK DISTRIBUTION BY OWNER ==='])
        writer.writerow(['Owner', 'Count'])
        for owner, count in stats['by_owner'].items():
            writer.writerow([owner, count])
        writer.writerow([])
        
        # Write effectiveness breakdown
        writer.writerow(['=== CONTROL EFFECTIVENESS ==='])
        writer.writerow(['Effectiveness', 'Count'])
        for eff, count in stats['by_effectiveness'].items():
            writer.writerow([eff, count])
        writer.writerow([])
        
        # Write mitigated risks
        writer.writerow(['=== MITIGATED RISKS ==='])
        writer.writerow(['Risk ID', 'Title', 'Score', 'Status', 'Owner', 'Category'])
        for risk in stats['mitigated_risks']:
            writer.writerow([
                risk['risk_id'],
                risk['title'],
                risk['risk_score'],
                risk['status'],
                risk['risk_owner'],
                risk['risk_category']
            ])
        writer.writerow([])
        
        # Write risks requiring attention
        writer.writerow(['=== RISKS REQUIRING ATTENTION ==='])
        writer.writerow(['Risk ID', 'Title', 'Score', 'Status', 'Owner', 'Category'])
        for risk in stats['not_mitigated_risks']:
            writer.writerow([
                risk['risk_id'],
                risk['title'],
                risk['risk_score'],
                risk['status'],
                risk['risk_owner'],
                risk['risk_category']
            ])
        writer.writerow([])
        
        # Write all risks
        writer.writerow(['=== ALL RISKS ==='])
        all_risks = list(Risk.objects.all().values(
            'risk_id', 'title', 'risk_owner', 'risk_category',
            'likelihood', 'impact', 'risk_score', 'status',
            'control_effectiveness', 'is_mitigated', 'last_updated'
        ))
        writer.writerow(['Risk ID', 'Title', 'Owner', 'Category', 'Likelihood', 'Impact', 'Score', 'Status', 'Effectiveness', 'Mitigated', 'Last Updated'])
        for risk in all_risks:
            writer.writerow([
                risk['risk_id'],
                risk['title'],
                risk['risk_owner'],
                risk['risk_category'],
                risk['likelihood'],
                risk['impact'],
                risk['risk_score'],
                risk['status'],
                risk['control_effectiveness'],
                'Yes' if risk['is_mitigated'] else 'No',
                str(risk['last_updated'])
            ])
        
        return output
    
    def update_persistent_csv(self):
        """Update the persistent CSV file."""
        import os
        from django.conf import settings
        
        # Ensure media directory exists
        media_root = settings.MEDIA_ROOT
        if not os.path.exists(media_root):
            os.makedirs(media_root)
            
        file_path = os.path.join(media_root, 'risk_management_report.csv')
        
        # Generate CSV content
        csv_buffer = self.generate_csv_report()
        
        # Write to file
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            f.write(csv_buffer.getvalue())
            
        return file_path

    def _create_plotly_risk_matrix(self, df):
        """Create a Plotly risk matrix figure."""
        if not PLOTLY_AVAILABLE or df.empty:
            return None
            
        # Group by likelihood and impact to get counts
        matrix_counts = df.groupby(['Likelihood', 'Impact']).size().reset_index(name='count')
        
        # Initialize 5x5 grid with zeros
        z_data = [[0 for _ in range(5)] for _ in range(5)]
        text_data = [['' for _ in range(5)] for _ in range(5)]
        
        # Fill grid
        for _, row in matrix_counts.iterrows():
            l = int(row['Likelihood']) - 1  # 0-indexed
            i = int(row['Impact']) - 1      # 0-indexed
            if 0 <= l < 5 and 0 <= i < 5:
                # Plotly heatmap uses (y, x) where y is likely likelihood (rows) and x is impact (cols)
                # But standard matrix has Likelihood on Y (1-5 ascending or descending?) 
                # Usually Likelihood is Y axis, Impact is X axis.
                # Let's align with common matrices: (0,0) is usually low/low.
                z_data[l][i] = row['count']
                text_data[l][i] = str(row['count'])

        # Create Heatmap
        fig = go.Figure(data=go.Heatmap(
            z=z_data,
            x=[1, 2, 3, 4, 5],
            y=[1, 2, 3, 4, 5],
            text=text_data,
            texttemplate="%{text}",
            textfont={"size": 16},
            colorscale=[
                [0, "#22c55e"],      # Green
                [0.25, "#eab308"],   # Yellow
                [0.5, "#f97316"],    # Orange
                [1, "#ef4444"]       # Red
            ],
            showscale=False
        ))

        fig.update_layout(
            title="Risk Matrix",
            xaxis_title="Impact",
            yaxis_title="Likelihood",
            width=800,
            height=600,
            template="plotly_white"
        )
        
        return fig

    def generate_excel_report(self, filters=None):
        """Generate comprehensive Excel report."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
            
        filters = filters or {}
        
        # 1. Prepare Data
        # Fetch all data as DataFrame
        risks_qs = Risk.objects.all().values(
            'risk_id', 'title', 'risk_owner', 'risk_category',
            'likelihood', 'impact', 'risk_score', 'status',
            'control_effectiveness', 'is_mitigated', 'last_updated'
        )
        
        if not risks_qs.exists():
            return b""
            
        df = pd.DataFrame(list(risks_qs))
        
        # Rename columns for friendly display
        df.rename(columns={
            'risk_id': 'Risk ID',
            'title': 'Title',
            'risk_owner': 'Owner',
            'risk_category': 'Category',
            'likelihood': 'Likelihood',
            'impact': 'Impact',
            'risk_score': 'Score',
            'status': 'Status',
            'control_effectiveness': 'Effectiveness',
            'is_mitigated': 'Is Mitigated',
            'last_updated': 'Last Updated'
        }, inplace=True)
        
        # Fix Is Mitigated display
        df['Is Mitigated'] = df['Is Mitigated'].apply(lambda x: 'Yes' if x else 'No')
        
        # 2. Prepare Risk Matrix Figure
        risk_matrix_fig = None
        if PLOTLY_AVAILABLE:
            try:
                risk_matrix_fig = self._create_plotly_risk_matrix(df)
            except Exception as e:
                print(f"Error creating Plotly figure: {e}")

        # 3. Generate Excel
        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # --- Sheet 1: Report Info ---
                stats = self.get_statistics()
                report_info = pd.DataFrame([
                    ["Report Generated", datetime.now().strftime('%Y-%m-%d %H:%M')],
                    ["Total Risks", len(df)],
                    ["Active Risks", len(df[~df['Status'].isin(['Mitigated', 'Closed'])])],
                    ["Mitigated Risks", len(df[df['Status'] == 'Mitigated'])],
                    ["Critical Risks", stats['by_severity'].get('Critical', 0)],
                    ["High Risks", stats['by_severity'].get('High', 0)],
                ] + [[f"Filter: {k}", str(v)] for k, v in filters.items()])
                
                report_info.to_excel(writer, index=False, header=False, sheet_name='Report_Info')
                
                # Styling Report Info
                ws_info = writer.sheets['Report_Info']
                ws_info.column_dimensions['A'].width = 25
                ws_info.column_dimensions['B'].width = 30
                for row in ws_info.iter_rows():
                    for cell in row:
                        cell.font = Font(size=12)
                        if cell.col_idx == 1:
                            cell.font = Font(bold=True, size=12, color='0D9488')
                
                # --- Sheet 2: All Risks ---
                df.to_excel(writer, index=False, sheet_name='All_Risks')
                
                # --- Sheet 3: Active Risks ---
                active_risks = df[~df['Status'].isin(['Mitigated', 'Closed'])]
                if not active_risks.empty:
                    active_risks.to_excel(writer, index=False, sheet_name='Active_Risks')
                
                # --- Sheet 4: Mitigated Risks ---
                mitigated_df = df[df['Status'] == 'Mitigated']
                if not mitigated_df.empty:
                    mitigated_df.to_excel(writer, index=False, sheet_name='Mitigated_Risks')
                
                # --- Sheet 5: Risk Matrix (Image) ---
                if risk_matrix_fig:
                    try:
                        # Create sheet
                        workbook = writer.book
                        matrix_ws = workbook.create_sheet("Risk_Matrix")
                        
                        # Generate image
                        img_bytes = pio.to_image(risk_matrix_fig, format="png", width=800, height=600, scale=2, engine="kaleido")
                        img = OpenpyxlImage(io.BytesIO(img_bytes))
                        
                        # Add image
                        matrix_ws.add_image(img, 'A1')
                        matrix_ws['A25'] = "Risk Matrix generated using Plotly"
                    except Exception as e:
                        print(f"Could not add risk matrix to Excel: {e}")
                        # If kaleido fails or other error, we just skip the image
                        pass

                # --- Formatting Loop (Auto-adjust widths & styles) ---
                header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                for sheet_name in writer.sheets:
                    if sheet_name == 'Report_Info' or sheet_name == 'Risk_Matrix':
                        continue
                        
                    worksheet = writer.sheets[sheet_name]
                    
                    # Format Header
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Auto-adjust columns
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                cell.border = border
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                        
            return output.getvalue()
            
        except Exception as e:
            print(f"Error generating Excel report: {e}")
            # Fallback or re-raise? simpler to return empty bytes or handle in view
            raise e
            eff_pie_chart.height = 10
            ws_summary.add_chart(eff_pie_chart, "D68")

        
        # Mitigated Risks Sheet
        ws_mitigated = wb.create_sheet('Mitigated Risks')
        mitigated_headers = ['Risk ID', 'Title', 'Score', 'Status', 'Owner', 'Category']
        
        for col, header in enumerate(mitigated_headers, 1):
            cell = ws_mitigated.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color='22C55E', end_color='22C55E', fill_type='solid')
            cell.font = header_font
            cell.border = border
        
        for row_idx, risk in enumerate(stats['mitigated_risks'], 2):
            ws_mitigated.cell(row=row_idx, column=1, value=risk['risk_id']).border = border
            ws_mitigated.cell(row=row_idx, column=2, value=risk['title']).border = border
            ws_mitigated.cell(row=row_idx, column=3, value=risk['risk_score']).border = border
            ws_mitigated.cell(row=row_idx, column=4, value=risk['status']).border = border
            ws_mitigated.cell(row=row_idx, column=5, value=risk['risk_owner']).border = border
            ws_mitigated.cell(row=row_idx, column=6, value=risk['risk_category']).border = border
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_mitigated.column_dimensions[col].width = 18 if col != 'B' else 35
        
        # Not Mitigated Risks Sheet
        ws_not_mitigated = wb.create_sheet('Risks Requiring Attention')
        
        for col, header in enumerate(mitigated_headers, 1):
            cell = ws_not_mitigated.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
            cell.font = header_font
            cell.border = border
        
        for row_idx, risk in enumerate(stats['not_mitigated_risks'], 2):
            ws_not_mitigated.cell(row=row_idx, column=1, value=risk['risk_id']).border = border
            ws_not_mitigated.cell(row=row_idx, column=2, value=risk['title']).border = border
            ws_not_mitigated.cell(row=row_idx, column=3, value=risk['risk_score']).border = border
            ws_not_mitigated.cell(row=row_idx, column=4, value=risk['status']).border = border
            ws_not_mitigated.cell(row=row_idx, column=5, value=risk['risk_owner']).border = border
            ws_not_mitigated.cell(row=row_idx, column=6, value=risk['risk_category']).border = border
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_not_mitigated.column_dimensions[col].width = 18 if col != 'B' else 35
        
        # All Risks Sheet
        ws_all = wb.create_sheet('All Risks')
        all_risks = list(Risk.objects.all().values(
            'risk_id', 'title', 'risk_owner', 'risk_category',
            'likelihood', 'impact', 'risk_score', 'status',
            'control_effectiveness', 'is_mitigated', 'last_updated'
        ))
        
        all_headers = ['Risk ID', 'Title', 'Owner', 'Category', 'Likelihood', 'Impact', 'Score', 'Status', 'Effectiveness', 'Mitigated', 'Last Updated']
        
        for col, header in enumerate(all_headers, 1):
            cell = ws_all.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        for row_idx, risk in enumerate(all_risks, 2):
            ws_all.cell(row=row_idx, column=1, value=risk['risk_id']).border = border
            ws_all.cell(row=row_idx, column=2, value=risk['title']).border = border
            ws_all.cell(row=row_idx, column=3, value=risk['risk_owner']).border = border
            ws_all.cell(row=row_idx, column=4, value=risk['risk_category']).border = border
            ws_all.cell(row=row_idx, column=5, value=risk['likelihood']).border = border
            ws_all.cell(row=row_idx, column=6, value=risk['impact']).border = border
            ws_all.cell(row=row_idx, column=7, value=risk['risk_score']).border = border
            ws_all.cell(row=row_idx, column=8, value=risk['status']).border = border
            ws_all.cell(row=row_idx, column=9, value=risk['control_effectiveness']).border = border
            ws_all.cell(row=row_idx, column=10, value='Yes' if risk['is_mitigated'] else 'No').border = border
            ws_all.cell(row=row_idx, column=11, value=str(risk['last_updated'])).border = border
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

